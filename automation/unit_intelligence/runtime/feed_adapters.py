#!/usr/bin/env python3
"""Feed adapters for approved portal data exports and CRM dumps.

These adapters read local files only (CSV, JSON, Excel). They do not scrape
portals or call external APIs. Use them after Omar has provided an authorized
export file.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from ingestion_queue import IngestionQueue


SUPPORTED_EXTENSIONS = {".csv", ".json", ".jsonl", ".xlsx", ".xls"}


def _load_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8", errors="replace") as fh:
        reader = csv.DictReader(fh)
        return [row for row in reader]


def _load_json(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return [data]
    return []


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    items = []
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                items.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return items


def _load_excel(path: Path) -> list[dict[str, Any]]:
    """Placeholder: requires openpyxl. We avoid adding dependencies until needed."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel files. Install with: pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({header: value for header, value in zip(headers, row) if header is not None})
    return rows


def _normalize_feed_row(row: dict[str, Any], portal: str) -> dict[str, Any]:
    """Normalize common column names into the canonical schema."""
    url = (
        row.get("listing_url")
        or row.get("url")
        or row.get("property_finder_url")
        or row.get("bayut_url")
        or row.get("dubizzle_url")
        or ""
    )
    listing_id = row.get("listing_id") or row.get("id") or row.get("reference") or ""
    bedrooms = row.get("bedrooms") or row.get("beds") or row.get("bedroom")
    try:
        bedrooms = int(bedrooms) if bedrooms is not None and str(bedrooms).strip() != "" else None
    except ValueError:
        bedrooms = None
    return {
        "portal": portal,
        "url": str(url).strip(),
        "listing_id": str(listing_id).strip() if listing_id else None,
        "feed_id": row.get("feed_id") or row.get("id") or row.get("reference") or None,
        "property_type": row.get("property_type") or row.get("type") or None,
        "area": row.get("area") or row.get("location") or row.get("community") or None,
        "project": row.get("project") or row.get("building_name") or None,
        "building": row.get("building") or row.get("tower") or None,
        "bedrooms": bedrooms,
        "price": row.get("price") or row.get("amount") or None,
        "size": row.get("size") or row.get("area_sqft") or row.get("sqft") or None,
        "developer": row.get("developer") or row.get("developer_name") or None,
        "permit_number": row.get("permit_number") or row.get("rera") or None,
        "property_number": row.get("property_number") or None,
        "plot_number": row.get("plot_number") or None,
    }


def load_feed_file(path: str | Path) -> list[dict[str, Any]]:
    """Load a feed file and return a list of normalized rows."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Feed file not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = _load_csv(path)
    elif suffix == ".json":
        raw = _load_json(path)
    elif suffix == ".jsonl":
        raw = _load_jsonl(path)
    elif suffix in {".xlsx", ".xls"}:
        raw = _load_excel(path)
    else:
        raise ValueError(f"Unsupported feed file extension: {suffix}")

    # Infer portal from filename or URL
    portal = "unknown"
    lowered = path.name.lower()
    if "propertyfinder" in lowered or "pf" in lowered:
        portal = "propertyfinder"
    elif "bayut" in lowered:
        portal = "bayut"
    elif "dubizzle" in lowered:
        portal = "dubizzle"

    rows = []
    for row in raw:
        normalized = _normalize_feed_row(row, portal)
        # Re-infer portal from URL if filename didn't give it away
        if normalized["portal"] == "unknown" and normalized["url"]:
            url_lower = normalized["url"].lower()
            if "propertyfinder" in url_lower:
                normalized["portal"] = "propertyfinder"
            elif "bayut" in url_lower:
                normalized["portal"] = "bayut"
            elif "dubizzle" in url_lower:
                normalized["portal"] = "dubizzle"
        rows.append(normalized)
    return rows


def ingest_feed_file(path: str | Path, queue: IngestionQueue | None = None) -> dict[str, Any]:
    """Load a feed file and ingest all rows into the staging database."""
    rows = load_feed_file(path)
    feed_items = []
    for row in rows:
        if not row.get("url"):
            continue
        feed_items.append(row)
    q = queue or IngestionQueue()
    ingested = q.ingest_portal_feed(feed_items)
    return {
        "ok": True,
        "file": str(path),
        "rows_loaded": len(rows),
        "rows_with_url": len(feed_items),
        "ingested": len(ingested),
    }


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        print(json.dumps(ingest_feed_file(sys.argv[1]), indent=2))
    else:
        print("Usage: feed_adapters.py <feed_file>")
