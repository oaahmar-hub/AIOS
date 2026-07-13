#!/usr/bin/env python3
"""Owner Lookup — unit / building / property-number -> real owner + phone.

Built on the DLD ownership exports (JVC.xlsx and the per-area files). This is
the keystone the Deal Agent uses at stage OWNERS: given a candidate unit it
returns the real owner name(s) + phone(s), or nothing (never a fabricated
number).

Handles the two DLD sheet schemas seen in the data:
  * "transaction" schema:  BuildingNameEn / UnitNumber / NameEn / Mobile /
    ProcedurePartyTypeNameEn (Seller|Buyer) / Plot Pre Reg No
  * "owner" schema:        BuildingName 2 / UnitNumber / property_number /
    Owner Name / Phone 1..2 / Mobile 1..2

Ingest writes a local SQLite index (``AIOS_OWNER_DB_PATH``, default under the
runtime data dir). The data is PII (names, mobiles, IDs) — it lives only in
that local/volume DB, is never committed, and is served ONLY behind admin auth
(the customer channel's NEVER_DISCLOSE guard stays intact).

Pure stdlib for query/health. Ingest uses openpyxl when available.
"""

from __future__ import annotations

import os
import re
import sqlite3
from pathlib import Path
from typing import Optional

_DEFAULT_DB = Path(os.getenv("AIOS_PHASE4_DB_PATH", "/tmp/x")).parent / "owner_lookup.sqlite"
OWNER_DB_PATH = Path(os.getenv("AIOS_OWNER_DB_PATH", str(_DEFAULT_DB)))

# Encrypted seed DB shipped in the repo (ciphertext only — never plaintext PII).
# On boot, if the live DB is empty and AIOS_OWNER_SEED_KEY is set, the server
# restores the real owner index from this seed. The key lives ONLY in the
# Railway env, so the git blob is useless without it.
_SEED_PATH = Path(__file__).resolve().parent / "data" / "owner_seed.db.gz.enc"
_SEED_RESTORED = False


def _seed_crypt(data: bytes, key: str, nonce: bytes) -> bytes:
    """Symmetric stream cipher (HMAC-SHA256 keystream, CTR mode) — stdlib only.
    XOR is its own inverse, so this both encrypts and decrypts."""
    import hashlib
    import hmac
    kb = hashlib.sha256(key.encode("utf-8")).digest()
    out = bytearray()
    i = 0
    while len(out) < len(data):
        out.extend(hmac.new(kb, nonce + i.to_bytes(8, "big"), hashlib.sha256).digest())
        i += 1
    return bytes(a ^ b for a, b in zip(data, out))


def _db_count_safe() -> int:
    try:
        con = sqlite3.connect(str(OWNER_DB_PATH))
        try:
            return con.execute("SELECT count(*) FROM owners").fetchone()[0]
        finally:
            con.close()
    except Exception:
        return 0


def _maybe_restore_seed() -> None:
    """Restore the real owner DB from the encrypted seed when the live DB is
    empty. Idempotent, best-effort, never raises into callers."""
    global _SEED_RESTORED
    if _SEED_RESTORED:
        return
    _SEED_RESTORED = True
    key = os.getenv("AIOS_OWNER_SEED_KEY", "")
    if not key or not _SEED_PATH.exists():
        return
    try:
        # Version marker = seed size; re-restore when the shipped seed changes
        # (e.g. JVC-only -> all-Dubai) even if a stale DB already sits on the
        # persistent volume.
        seed_sig = str(_SEED_PATH.stat().st_size)
        marker = OWNER_DB_PATH.with_name(OWNER_DB_PATH.name + ".seedver")
        if OWNER_DB_PATH.exists() and _db_count_safe() > 0:
            try:
                if marker.exists() and marker.read_text().strip() == seed_sig:
                    return  # DB already matches this seed
            except Exception:
                return
        import gzip
        blob = _SEED_PATH.read_bytes()
        nonce, ct = blob[:16], blob[16:]
        raw = gzip.decompress(_seed_crypt(ct, key, nonce))
        OWNER_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        OWNER_DB_PATH.write_bytes(raw)
        try:
            marker.write_text(seed_sig)
        except Exception:
            pass
    except Exception:
        pass

_OWNER_COLS = ["NameEn", "Owner Name", "OwnerName"]
_PHONE_COLS = ["Mobile 1", "Mobile 2", "Phone 1", "Phone 2", "Mobile", "Phone"]

# Column aliases (matched case-insensitively) covering every DLD/developer export
# format seen in Omar's data: transaction sheets, plot "P-NUMBER" exports (owner +
# property split across two sheets, joined on P-NUMBER), single "P & O" sheets, and
# developer master files. This is what lets ONE ingest swallow every area file.
_ALIAS = {
    "name":  ["nameen", "owner name", "ownername", "name", "primary applicant name", "owner"],
    "phone": ["mobile 1", "mobile 2", "mobile", "primary mobile number", "secondary mobile",
              "mobile number", "phone 1", "phone 2", "phone"],
    "building": ["buildingname 2", "buildingnameen", "building name", "tower name", "tower",
                 "building 1", "building no", "sub project", "master project", "project"],
    "unit": ["unitnumber", "unit number", "unit", "flat number"],
    "pnum": ["property_number", "p-number", "pnumber", "plot pre reg no",
             "registration number", "municipality number"],
    "plot": ["plot number", "land number", "landnumber"],
    "area": ["area", "master location"],
    "country": ["countrynameen", "residence country", "nationality"],
    "role": ["procedurepartytypenameen"],
}


def _hmap(header: list) -> dict:
    """lowercased-header -> column index."""
    return {_norm(h).lower(): i for i, h in enumerate(header) if h is not None}


def _col(hm: dict, key: str):
    for a in _ALIAS[key]:
        if a in hm:
            return hm[a]
    return None


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def _norm(x) -> str:
    return str(x).strip() if x is not None else ""


def _clean_phone(s: str) -> str:
    """Keep a phone only if it has a plausible number of digits."""
    s = _norm(s)
    d = "".join(c for c in s if c.isdigit())
    if len(d) < 7 or d.strip("0") == "":
        return ""
    return s


def _first_phone(*vals) -> str:
    for v in vals:
        p = _clean_phone(v)
        if p:
            return p
    return ""


def _clean(v: str) -> str:
    """Blank out junk placeholders ('null'/'none'/'nan') from DLD cells."""
    s = _norm(v)
    return "" if s.lower() in ("null", "none", "nan", "n/a", "-") else s


def mask(phone: str) -> str:
    d = "".join(c for c in _norm(phone) if c.isdigit())
    return f"***{d[-3:]}" if len(d) >= 3 else "***"


def _connect() -> sqlite3.Connection:
    _maybe_restore_seed()
    OWNER_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(str(OWNER_DB_PATH))
    con.row_factory = sqlite3.Row
    return con


def _ensure_schema(con: sqlite3.Connection) -> None:
    con.execute(
        """CREATE TABLE IF NOT EXISTS owners(
            area TEXT, building TEXT, unit TEXT, project TEXT,
            property_number TEXT, role TEXT, name TEXT, phone TEXT,
            country TEXT, source TEXT)"""
    )
    con.execute("CREATE INDEX IF NOT EXISTS ix_bu ON owners(building, unit)")
    con.execute("CREATE INDEX IF NOT EXISTS ix_pn ON owners(property_number)")
    con.execute("CREATE INDEX IF NOT EXISTS ix_unit ON owners(unit)")


# ---------------------------------------------------------------------------
# ingest
# ---------------------------------------------------------------------------
def index_rows(rows: list, source: str = "manual", reset: bool = False) -> int:
    """Insert normalized owner rows. Each row is a dict with any of:
    area, building, unit, project, property_number, role, name, phone, country.
    Returns the number inserted."""
    con = _connect()
    try:
        _ensure_schema(con)
        if reset:
            con.execute("DELETE FROM owners WHERE source = ?", (source,))
        buf = []
        for r in rows:
            name = _norm(r.get("name"))
            if not name:
                continue
            buf.append((
                _norm(r.get("area")), _norm(r.get("building")), _norm(r.get("unit")),
                _norm(r.get("project")), _norm(r.get("property_number")),
                _norm(r.get("role")) or "Owner", name, _clean_phone(r.get("phone")),
                _norm(r.get("country")), source,
            ))
        con.executemany(
            "INSERT INTO owners(area,building,unit,project,property_number,role,name,phone,country,source)"
            " VALUES(?,?,?,?,?,?,?,?,?,?)", buf,
        )
        con.commit()
        return len(buf)
    finally:
        con.close()


def ingest_dld_xlsx(path: str, area: str = "") -> dict:
    """Parse ANY DLD/developer xlsx (all known formats) into the index.

    Handles: transaction sheets (BuildingNameEn/NameEn/Mobile), owner sheets
    (Owner Name/property_number/Phone), plot "P-NUMBER" exports where owner and
    property live in SEPARATE sheets (joined on P-NUMBER), single "P & O" sheets,
    and developer master files (Tower/Unit/Applicant/Mobile). Column names are
    matched case-insensitively via _ALIAS, so one code path swallows every file.
    """
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover
        return {"ok": False, "error": f"openpyxl unavailable: {exc}"}
    p = Path(path)
    area = area or re.sub(r"[^A-Za-z0-9 ]", "", p.stem).strip()
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}

    # Read every sheet once: (name, header-map, list-of-rows).
    sheets = []
    for sheet in wb.sheetnames:
        it = wb[sheet].iter_rows(values_only=True)
        try:
            hm = _hmap(list(next(it)))
        except StopIteration:
            continue
        sheets.append((sheet, hm, list(it)))

    def gv(row, idx):
        return _norm(row[idx]) if idx is not None and idx < len(row) else ""

    # Pass 1: build a property lookup keyed by P-NUMBER / plot, from any sheet
    # that carries building/unit — so owner rows lacking building/unit can join.
    prop_by_key: dict = {}
    for _name, hm, rows in sheets:
        b, u, pn, pl = _col(hm, "building"), _col(hm, "unit"), _col(hm, "pnum"), _col(hm, "plot")
        if (b is None and u is None) or (pn is None and pl is None):
            continue
        for r in rows:
            bldg, unit = gv(r, b), gv(r, u)
            if not bldg and not unit:
                continue
            for k in (gv(r, pn), gv(r, pl)):
                if k and k not in prop_by_key:
                    prop_by_key[k] = (bldg, unit)

    # Pass 2: emit owner rows from any sheet that has a name column.
    total = 0
    for _name, hm, rows in sheets:
        ni = _col(hm, "name")
        if ni is None:
            continue
        bi, ui, pni, pli = _col(hm, "building"), _col(hm, "unit"), _col(hm, "pnum"), _col(hm, "plot")
        ri, ci = _col(hm, "role"), _col(hm, "country")
        phone_idx = [hm[a] for a in _ALIAS["phone"] if a in hm]
        # Community / project column — villas (Golf Place, etc.) have no building
        # name; they're identified by their community, so capture it and use it
        # as the searchable "building" when BuildingNameEn is blank.
        proj_i = None
        for cand in ("project", "project name en", "project name", "project_name_en",
                     "master project", "master_project_en", "master project en", "community"):
            if cand in hm:
                proj_i = hm[cand]; break
        out = []
        for r in rows:
            name = gv(r, ni)
            if not name:
                continue
            bldg, unit = gv(r, bi), gv(r, ui)
            proj = gv(r, proj_i) if proj_i is not None else ""
            pnum = gv(r, pni)
            if not bldg or not unit:  # join from the property sheet
                for k in (pnum, gv(r, pli)):
                    if k and k in prop_by_key:
                        jb, ju = prop_by_key[k]
                        bldg = bldg or jb
                        unit = unit or ju
                        break
            # Villa/community fallback: no real building -> make the community the
            # searchable/displayed building (villas carry "null" or blank here).
            if (not bldg or str(bldg).strip().lower() in ("null", "none")) and proj:
                bldg = proj
            phone = _first_phone(*[gv(r, i) for i in phone_idx])
            out.append({
                "area": area, "building": bldg, "unit": unit, "project": proj,
                "property_number": pnum, "role": gv(r, ri) or "Owner",
                "name": name, "phone": phone, "country": gv(r, ci),
            })
        total += index_rows(out, source=f"dld:{p.name}:{_name}")
    return {"ok": True, "file": p.name, "area": area, "indexed": total}


# ---------------------------------------------------------------------------
# lookup
# ---------------------------------------------------------------------------
def lookup(building: str = "", unit: str = "", property_number: str = "",
           area: str = "", q: str = "", limit: int = 10, reveal: bool = False) -> dict:
    """Return owners matching the given keys. Phones masked unless reveal=True
    (reveal is only ever passed after admin auth at the API layer).

    `q` is a free search matching building OR area (per token), so typing an
    area name ("Meydan", "Business Bay") finds owners there, not just buildings.
    """
    try:
        if not any([building.strip(), unit.strip(), property_number.strip(), q.strip(), area.strip()]):
            return {"ok": False, "error": "need building, area, unit, or property_number"}
        con = _connect()
        try:
            _ensure_schema(con)
            clauses, params = [], []
            _filler = {"the", "of", "at", "by", "and", "dubai"}
            if q.strip():
                # each token must appear in building OR area OR project/community
                for t in [t for t in re.split(r"[^a-z0-9]+", q.strip().lower())
                          if len(t) > 1 and t not in _filler]:
                    clauses.append("(lower(building) LIKE ? OR lower(area) LIKE ? OR lower(project) LIKE ?)")
                    params += [f"%{t}%", f"%{t}%", f"%{t}%"]
            if property_number.strip():
                clauses.append("lower(property_number) = ?"); params.append(property_number.strip().lower())
            if building.strip():
                # token-AND match: every significant word must appear in the
                # building, in any order — so "Bloom Towers" (or a URL-derived
                # "bloom towers") matches "BLOOM TOWERS C". Generic filler words
                # are dropped so they don't over-constrain.
                _filler = {"the", "of", "at", "by", "and", "dubai"}
                toks = [t for t in re.split(r"[^a-z0-9]+", building.strip().lower())
                        if len(t) > 1]
                core = [t for t in toks if t not in _filler] or toks
                for t in core:
                    clauses.append("(lower(building) LIKE ? OR lower(project) LIKE ?)")
                    params += [f"%{t}%", f"%{t}%"]
            if unit.strip():
                clauses.append("lower(unit) = ?"); params.append(unit.strip().lower())
            if area.strip():
                clauses.append("lower(area) LIKE ?"); params.append(f"%{area.strip().lower()}%")
            where = " AND ".join(clauses)
            rows = con.execute(
                f"SELECT area,building,unit,property_number,role,name,phone,country "
                f"FROM owners WHERE {where} AND phone != '' LIMIT ?",
                params + [max(1, min(limit, 50))],
            ).fetchall()
        finally:
            con.close()
        out = []
        for r in rows:
            out.append({
                "area": _clean(r["area"]), "building": _clean(r["building"]), "unit": _clean(r["unit"]),
                "property_number": _clean(r["property_number"]), "role": r["role"],
                "name": r["name"], "country": _clean(r["country"]),
                "phone": r["phone"] if reveal else mask(r["phone"]),
            })
        return {"ok": True, "matches": len(out), "owners": out}
    except Exception as exc:  # pragma: no cover
        return {"ok": False, "error": str(exc)}


def search_units(query: str = "", area: str = "", building: str = "", limit: int = 25) -> list:
    """Find DLD-registered UNITS (distinct building/unit/area) across every
    ingested area — this turns the owner index into a Dubai-wide unit finder.
    Never returns phones; use lookup() for owner contact."""
    try:
        con = _connect()
        try:
            _ensure_schema(con)
            clauses, params = [], []
            if area.strip():
                clauses.append("lower(area) LIKE ?"); params.append(f"%{area.strip().lower()}%")
            if building.strip():
                for t in [t for t in re.split(r"[^a-z0-9]+", building.lower()) if len(t) > 1]:
                    clauses.append("lower(building) LIKE ?"); params.append(f"%{t}%")
            if query.strip():
                for t in [t for t in re.split(r"[^a-z0-9]+", query.lower()) if len(t) > 2]:
                    clauses.append("(lower(building) LIKE ? OR lower(area) LIKE ?)")
                    params += [f"%{t}%", f"%{t}%"]
            where = " AND ".join(clauses) if clauses else "1=1"
            rows = con.execute(
                f"SELECT DISTINCT area,building,unit,property_number FROM owners "
                f"WHERE {where} AND building != '' LIMIT ?",
                params + [max(1, min(limit, 100))],
            ).fetchall()
        finally:
            con.close()
        return [{
            "area": _clean(r["area"]), "building": _clean(r["building"]), "unit": _clean(r["unit"]),
            "property_number": _clean(r["property_number"]), "source": "DLD registered",
        } for r in rows]
    except Exception:
        return []


def owners_for_unit(unit: dict) -> list:
    """Deal Agent adapter: candidate unit dict -> [{name, phone(real)}]. Server
    side only (reveal=True) — the outbound guardrails live in the Deal Agent."""
    res = lookup(
        building=str(unit.get("building", "")),
        unit=str(unit.get("unit", "")),
        property_number=str(unit.get("property_number", "")),
        reveal=True, limit=5,
    )
    return res.get("owners", []) if res.get("ok") else []


def health() -> dict:
    try:
        con = _connect()
        try:
            _ensure_schema(con)
            n = con.execute("SELECT count(*) FROM owners").fetchone()[0]
            withph = con.execute("SELECT count(*) FROM owners WHERE phone != ''").fetchone()[0]
            areas = con.execute("SELECT count(DISTINCT area) FROM owners").fetchone()[0]
        finally:
            con.close()
        return {"component": "owner_lookup", "records": int(n), "with_phone": int(withph),
                "areas": int(areas), "db": str(OWNER_DB_PATH),
                "status": "ok" if n else "empty_awaiting_ingest"}
    except Exception as exc:
        return {"component": "owner_lookup", "status": f"error:{exc}"}
