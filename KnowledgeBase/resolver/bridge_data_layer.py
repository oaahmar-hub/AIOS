#!/usr/bin/env python3
"""AIOS bridge data layer: import, normalize, classify, and look up bridge rows."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List


RESOLVER_DIR = Path(__file__).resolve().parent


def resolve_resolver_db(resolver_dir: Path) -> Path:
    """Return the resolver database path, preferring the ``.resolver`` file."""
    resolver_file = resolver_dir / "unit_resolver_database.resolver"
    if resolver_file.is_file():
        return resolver_file
    return resolver_dir / "unit_resolver_database.sqlite"


DB_PATH = resolve_resolver_db(RESOLVER_DIR)
LISTING_IDENTITY_JSON = RESOLVER_DIR / "listing_identity_map.json"
LISTING_IDENTITY_CSV = RESOLVER_DIR / "listing_identity_map.csv"
INDEX_CSV = RESOLVER_DIR / "unit_resolver_index.csv"
BRIDGE_EXPORT_CSV = RESOLVER_DIR / "bridge_records_export.csv"
BRIDGE_REPORT_MD = RESOLVER_DIR / "bridge_import_report.md"
TABLE_NAME = "bridge_records"

URL_RE = re.compile(r"(?:https?://)?(?:www\.)?(?:propertyfinder\.ae|bayut\.com|dubizzle\.com)[^\s\"'<>)\]]+", re.I)
LISTING_ID_RE = re.compile(r"(?:^|[/?#=&_-])([0-9]{5,20})(?:$|[/?#=&_.-])")
PROPERTY_TYPES = {"apartment", "villa", "studio", "office", "warehouse", "shop", "commercial", "townhouse", "land"}
FIELD_ALIASES = {
    "source_platform": ["source_platform", "platform", "domain"],
    "listing_url": ["listing_url", "url", "property_finder_url", "bayut_url", "dubizzle_url"],
    "listing_id": ["listing_id", "listing_reference", "listing_ref", "listingid"],
    "broker_reference": ["broker_reference", "broker_ref", "brn", "orn", "reference", "ref"],
    "permit_number": ["permit_number", "permit", "trakheesi", "dld_permit"],
    "property_number": ["property_number", "property_no", "property_id"],
    "plot_number": ["plot_number", "plot_no", "plot"],
    "building_name": ["building_name", "building", "tower"],
    "unit_number": ["unit_number", "unit", "unit_ref"],
    "canonical_property_id": ["canonical_property_id", "cpid", "canonical_id"],
    "confidence": ["confidence", "confidence_score"],
    "source_updated_at": ["source_updated_at", "updated_at", "last_modified", "file_date"],
    "created_at": ["created_at", "imported_at"],
    "community": ["community", "area"],
    "project_name": ["project_name", "project"],
    "size_clues": ["size_clues", "size"],
    "price_clues": ["price_clues", "price"],
    "resolver_record_id": ["resolver_record_id"],
    "owner_contact_available": ["owner_contact_available"],
    "source_file": ["source_file", "file_name"],
    "source_sheet": ["source_sheet", "sheet_name"],
    "row_number": ["row_number"],
}

SCHEMA_SQL = f"""
CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
    bridge_id TEXT PRIMARY KEY,
    source_platform TEXT,
    listing_url TEXT,
    normalized_url TEXT,
    listing_id TEXT,
    broker_reference TEXT,
    permit_number TEXT,
    property_number TEXT,
    plot_number TEXT,
    building_name TEXT,
    unit_number TEXT,
    community TEXT,
    project_name TEXT,
    canonical_property_id TEXT,
    resolver_record_id TEXT,
    bridge_classification TEXT,
    confidence INTEGER,
    slug_clues TEXT,
    size_clues TEXT,
    price_clues TEXT,
    owner_contact_available TEXT,
    source_updated_at TEXT,
    created_at TEXT,
    source_file TEXT,
    source_sheet TEXT,
    row_number TEXT,
    raw_payload_json TEXT
);
CREATE INDEX IF NOT EXISTS idx_bridge_url ON {TABLE_NAME}(normalized_url);
CREATE INDEX IF NOT EXISTS idx_bridge_listing_id ON {TABLE_NAME}(listing_id);
CREATE INDEX IF NOT EXISTS idx_bridge_broker_reference ON {TABLE_NAME}(broker_reference);
CREATE INDEX IF NOT EXISTS idx_bridge_property_number ON {TABLE_NAME}(property_number);
CREATE INDEX IF NOT EXISTS idx_bridge_permit_number ON {TABLE_NAME}(permit_number);
CREATE INDEX IF NOT EXISTS idx_bridge_plot_number ON {TABLE_NAME}(plot_number);
CREATE INDEX IF NOT EXISTS idx_bridge_cpid ON {TABLE_NAME}(canonical_property_id);
"""


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def low(value: object) -> str:
    return norm(value).lower()


def key(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", low(value))


# Words that reveal a value is search-slug / listing prose rather than a real
# property identifier (e.g. "for sale dubai dubai land majan 1583088").
SLUG_JUNK_TOKENS = (
    "for sale",
    "for rent",
    "market analysis",
    "tru estimate",
    "awards",
    "blog",
    "agent",
    "listed",
    "details",
    "finder",
    "dubai land",
    "majan",
)

# Single-token values that are placeholders rather than identifiers.
JUNK_IDENTIFIER_VALUES = {"details", "listed", "finder", "scaped", "-", "n/a", "na", "none", "null"}

# Building names known to be corrupt placeholders injected by the broken parser.
PLACEHOLDER_BUILDINGS = {"canal bay"}

# URL path segments that indicate an auth/redirect link rather than a listing.
NON_LISTING_URL_MARKERS = (
    "/auth/",
    "auth.bayut.com",
    "passwordless-login",
    "action-token",
    "magic-link",
    "/signin",
    "/login",
    "/callback",
    "/go/",
    "access_token",
    "id_token",
)


def clean_identifier(value: str) -> str:
    """Return a trimmed identifier, or "" if the value is slug/URL/junk text."""
    text = norm(value)
    if not text:
        return ""
    lowered = low(text)
    if lowered.startswith("/") or "propertyfinder" in lowered or "bayut" in lowered or "dubizzle" in lowered:
        return ""
    if "/" in text or "\\" in text:
        return ""
    if lowered in JUNK_IDENTIFIER_VALUES:
        return ""
    if any(token in lowered for token in SLUG_JUNK_TOKENS):
        return ""
    # Real identifiers are compact; slug prose has many words.
    if len(text.split()) > 3:
        return ""
    return text


def is_placeholder_building(value: str) -> bool:
    return low(value) in PLACEHOLDER_BUILDINGS


def is_valid_listing_id(value: str) -> bool:
    """True for compact portal listing ids, not JWT/auth-token fragments."""
    text = norm(value)
    if not text or len(text) > 24:
        return False
    lowered = low(text)
    if lowered.startswith(("eyj", "action-token", "action_token")) or "token" in lowered:
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9]{4,24}", text))


def is_listing_url(url: str) -> bool:
    """True only for real portal listing URLs (not auth/redirect/login links)."""
    lowered = low(url)
    if not lowered:
        return False
    if not any(portal in lowered for portal in ("propertyfinder.ae", "bayut.com", "dubizzle.com")):
        return False
    return not any(marker in lowered for marker in NON_LISTING_URL_MARKERS)


def is_strong_identifier(value: str, field_name: str) -> bool:
    text = clean_identifier(value)
    if not text:
        return False
    if len(text) > 40:
        return False
    # A hard identifier must contain a run of digits; pure words are locations.
    if not re.search(r"\d{2,}", text):
        return False
    if field_name in {"permit_number", "plot_number"}:
        return bool(re.fullmatch(r"[A-Za-z0-9-]{3,30}", text))
    if field_name == "property_number":
        return bool(re.fullmatch(r"[A-Za-z0-9 ]{3,40}", text))
    if field_name == "unit_number":
        return bool(re.fullmatch(r"[A-Za-z0-9 -]{1,20}", text))
    return False


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def clean_url(url: str) -> str:
    value = norm(url).rstrip(".,;")
    if value and not value.lower().startswith(("http://", "https://")):
        value = "https://" + value
    return value


def normalize_public_url(url: str) -> Dict[str, str]:
    original = clean_url(url)
    if not original:
        return {
            "source_platform": "",
            "listing_url": "",
            "normalized_url": "",
            "listing_id": "",
            "slug_clues": "",
            "community": "",
            "project_name": "",
            "building_name": "",
            "size_clues": "",
            "price_clues": "",
        }
    normalized = re.sub(r"[?#].*$", "", original).rstrip("/")
    platform = ""
    if "propertyfinder.ae" in normalized.lower():
        platform = "propertyfinder.ae"
    elif "bayut.com" in normalized.lower():
        platform = "bayut.com"
    elif "dubizzle.com" in normalized.lower():
        platform = "dubizzle.com"
    tail = normalized.split("/")[-1]
    listing_id = ""
    for match in reversed(LISTING_ID_RE.findall(tail)):
        if match.isdigit():
            listing_id = match
            break
    if not listing_id:
        digits = LISTING_ID_RE.findall(normalized)
        listing_id = digits[-1] if digits else ""
    slug = re.sub(r"\.(html?|htm)$", "", tail, flags=re.I)
    slug_tokens = [t for t in re.split(r"[^a-zA-Z0-9]+", slug) if t and not t.isdigit()]
    community = ""
    project_name = ""
    building_name = ""
    if platform == "propertyfinder.ae":
        if "dubai-south" in normalized.lower():
            community = "Dubai South"
        if "dubai-world-central" in normalized.lower():
            project_name = "Dubai World Central"
        if "dubai-logistics-city" in normalized.lower():
            building_name = "Dubai Logistics City"
    slug_clues = " ".join(slug_tokens[:20])
    return {
        "source_platform": platform,
        "listing_url": original,
        "normalized_url": normalized.lower(),
        "listing_id": listing_id,
        "slug_clues": slug_clues,
        "community": community,
        "project_name": project_name,
        "building_name": building_name,
        "size_clues": "",
        "price_clues": "",
    }


def canonical_property_id_from_row(row: Dict[str, str]) -> str:
    explicit = norm(row.get("canonical_property_id", ""))
    if explicit:
        return explicit
    resolver_record_id = norm(row.get("resolver_record_id", ""))
    if resolver_record_id:
        return f"CPID-{resolver_record_id}"
    for field, prefix in [("property_number", "PROP"), ("permit_number", "PERMIT"), ("plot_number", "PLOT"), ("unit_number", "UNIT")]:
        value = key(row.get(field, ""))
        if value:
            return f"CPID-{prefix}-{value}"
    seed = "|".join(
        [
            low(row.get("source_platform", "")),
            low(row.get("listing_id", "")),
            low(row.get("broker_reference", "")),
            low(row.get("building_name", "")),
            low(row.get("unit_number", "")),
        ]
    )
    return "CPID-HASH-" + hashlib.sha1(seed.encode("utf-8")).hexdigest()[:16]


def classify_bridge_row(row: Dict[str, str]) -> Dict[str, str]:
    has_public_ref = (
        is_listing_url(row.get("listing_url", ""))
        or is_valid_listing_id(row.get("listing_id", ""))
        or bool(norm(row.get("broker_reference", "")))
    )
    has_hard_property = any(is_strong_identifier(row.get(field, ""), field) for field in ("property_number", "permit_number", "plot_number", "unit_number"))
    has_location = any(
        norm(row.get(field, "")) and not is_placeholder_building(row.get(field, ""))
        for field in ("community", "project_name", "building_name")
    )
    cpid = canonical_property_id_from_row(row)
    if has_public_ref and has_hard_property and cpid:
        bridge_classification = "exact_bridge"
        confidence = 100
    elif has_public_ref and (has_location or cpid):
        bridge_classification = "partial_bridge"
        confidence = 80
    elif has_public_ref:
        bridge_classification = "candidate_bridge"
        confidence = 65
    else:
        bridge_classification = "invalid_bridge"
        confidence = 0
    row["canonical_property_id"] = cpid
    row["bridge_classification"] = bridge_classification
    row["confidence"] = str(confidence)
    return row


def bridge_id_for_row(row: Dict[str, str]) -> str:
    strong = [
        low(row.get("source_platform", "")),
        low(row.get("normalized_url", "")),
        low(row.get("listing_id", "")),
        low(row.get("broker_reference", "")),
        low(row.get("canonical_property_id", "")),
        low(row.get("resolver_record_id", "")),
    ]
    return hashlib.sha1("|".join(strong).encode("utf-8")).hexdigest()[:24]


def normalize_bridge_row(raw_row: Dict[str, object]) -> Dict[str, str]:
    row = {field: "" for field in FIELD_ALIASES}
    lowered = {low(k): v for k, v in raw_row.items()}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in lowered and norm(lowered[alias]):
                row[field] = norm(lowered[alias])
                break
    url_fields = [row.get("listing_url", "")]
    for alias in ("property_finder_url", "bayut_url", "dubizzle_url", "url"):
        value = norm(lowered.get(alias, ""))
        if value:
            url_fields.append(value)
    best_url = next((value for value in url_fields if value), "")
    normalized_url_data = normalize_public_url(best_url)
    row["source_platform"] = row["source_platform"] or normalized_url_data["source_platform"]
    row["listing_url"] = row["listing_url"] or normalized_url_data["listing_url"]
    row["normalized_url"] = normalized_url_data["normalized_url"]
    row["listing_id"] = row["listing_id"] or normalized_url_data["listing_id"]
    row["broker_reference"] = clean_identifier(row.get("broker_reference", ""))
    row["permit_number"] = clean_identifier(row.get("permit_number", ""))
    row["property_number"] = clean_identifier(row.get("property_number", ""))
    row["plot_number"] = clean_identifier(row.get("plot_number", ""))
    row["unit_number"] = clean_identifier(row.get("unit_number", ""))
    if row["listing_url"] and not is_listing_url(row["listing_url"]):
        row["listing_url"] = ""
        row["normalized_url"] = ""
    if row["listing_id"] and not is_valid_listing_id(row["listing_id"]):
        row["listing_id"] = ""
    row["community"] = row["community"] or normalized_url_data["community"]
    row["project_name"] = row["project_name"] or normalized_url_data["project_name"]
    row["building_name"] = row["building_name"] or normalized_url_data["building_name"]
    if is_placeholder_building(row["building_name"]):
        row["building_name"] = ""
    row["slug_clues"] = normalized_url_data["slug_clues"]
    row["created_at"] = row["created_at"] or now_iso()
    row["source_updated_at"] = row["source_updated_at"] or ""
    row["owner_contact_available"] = "YES" if low(row.get("owner_contact_available", "")) == "yes" else "NO"
    row["raw_payload_json"] = json.dumps(raw_row, ensure_ascii=False)
    return classify_bridge_row(row)


def iter_csv_rows(path: Path) -> Iterable[Dict[str, object]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            yield dict(row)


def iter_json_rows(path: Path) -> Iterable[Dict[str, object]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        for row in data:
            if isinstance(row, dict):
                yield row
    elif isinstance(data, dict):
        for row in data.get("rows", []):
            if isinstance(row, dict):
                yield row


def iter_xlsx_rows(path: Path) -> Iterable[Dict[str, object]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [low(cell) for cell in next(rows)]
        except StopIteration:
            continue
        for raw_values in rows:
            row = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = raw_values[idx] if idx < len(raw_values) else ""
                row[header] = value
            if any(norm(v) for v in row.values()):
                row.setdefault("source_sheet", ws.title)
                yield row


def load_rows_from_path(path: Path) -> List[Dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return list(iter_csv_rows(path))
    if suffix == ".json":
        return list(iter_json_rows(path))
    if suffix in {".xlsx", ".xlsm"}:
        return list(iter_xlsx_rows(path))
    return []


def ensure_bridge_table() -> None:
    con = sqlite3.connect(DB_PATH)
    try:
        con.executescript(SCHEMA_SQL)
        con.commit()
    finally:
        con.close()


def import_bridge_rows(paths: Iterable[Path], reset: bool = False) -> Dict[str, int]:
    ensure_bridge_table()
    rows: Dict[str, Dict[str, str]] = {}
    for path in paths:
        if not path.exists():
            continue
        for raw_row in load_rows_from_path(path):
            normalized = normalize_bridge_row(raw_row)
            bridge_id = bridge_id_for_row(normalized)
            normalized["bridge_id"] = bridge_id
            rows[bridge_id] = normalized
    con = sqlite3.connect(DB_PATH)
    try:
        if reset:
            con.execute(f"DELETE FROM {TABLE_NAME}")
        sql = f"""
        INSERT OR REPLACE INTO {TABLE_NAME} (
            bridge_id, source_platform, listing_url, normalized_url, listing_id, broker_reference,
            permit_number, property_number, plot_number, building_name, unit_number, community,
            project_name, canonical_property_id, resolver_record_id, bridge_classification, confidence,
            slug_clues, size_clues, price_clues, owner_contact_available, source_updated_at, created_at,
            source_file, source_sheet, row_number, raw_payload_json
        ) VALUES (
            :bridge_id, :source_platform, :listing_url, :normalized_url, :listing_id, :broker_reference,
            :permit_number, :property_number, :plot_number, :building_name, :unit_number, :community,
            :project_name, :canonical_property_id, :resolver_record_id, :bridge_classification, :confidence,
            :slug_clues, :size_clues, :price_clues, :owner_contact_available, :source_updated_at, :created_at,
            :source_file, :source_sheet, :row_number, :raw_payload_json
        )
        """
        con.executemany(sql, list(rows.values()))
        con.commit()
        counts = Counter(row["bridge_classification"] for row in rows.values())
        counts["imported_rows"] = len(rows)
        return dict(counts)
    finally:
        con.close()


def seed_default_bridge_rows(reset: bool = True) -> Dict[str, int]:
    sources = []
    if LISTING_IDENTITY_JSON.exists():
        sources.append(LISTING_IDENTITY_JSON)
    elif LISTING_IDENTITY_CSV.exists():
        sources.append(LISTING_IDENTITY_CSV)
    if INDEX_CSV.exists():
        sources.append(INDEX_CSV)
    counts = import_bridge_rows(sources, reset=reset)
    export_bridge_csv()
    write_bridge_report(counts)
    return counts


def load_bridge_rows() -> List[Dict[str, str]]:
    ensure_bridge_table()
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        return [dict(row) for row in con.execute(f"SELECT * FROM {TABLE_NAME}")]
    finally:
        con.close()


def export_bridge_csv() -> None:
    rows = load_bridge_rows()
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    with BRIDGE_EXPORT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_bridge_report(counts: Dict[str, int]) -> None:
    lines = [
        "# Bridge Import Report",
        "",
        f"- imported_rows: {counts.get('imported_rows', 0)}",
        f"- exact_bridge: {counts.get('exact_bridge', 0)}",
        f"- partial_bridge: {counts.get('partial_bridge', 0)}",
        f"- candidate_bridge: {counts.get('candidate_bridge', 0)}",
        f"- invalid_bridge: {counts.get('invalid_bridge', 0)}",
        f"- export_csv: `{BRIDGE_EXPORT_CSV}`",
    ]
    BRIDGE_REPORT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def bridge_lookup(field: str, value: str) -> List[Dict[str, str]]:
    target = key(value if field not in {"listing_url"} else normalize_public_url(value)["normalized_url"])
    if not target:
        return []
    rows = []
    for row in load_bridge_rows():
        source_value = row.get(field, "")
        compare = row.get("normalized_url", "") if field == "listing_url" else source_value
        if key(compare) == target:
            rows.append(row)
    rows.sort(key=lambda row: (int(row.get("confidence", "0") or 0), row.get("bridge_classification", "")), reverse=True)
    return rows


if __name__ == "__main__":
    counts = seed_default_bridge_rows(reset=True)
    print(json.dumps(counts, ensure_ascii=False, indent=2))
