#!/usr/bin/env python3
"""Import external CRM / broker / portal bridge datasets into AIOS PropertyGraph.

Accepted inputs:
- CSV
- XLSX
- JSON (list of dict rows)

Normalized fields:
- source_platform
- listing_url
- listing_id
- broker_reference
- permit_number
- property_number
- plot_number
- building_name
- unit_number
- canonical_property_id
- resolver_record_id
- confidence
- source_updated_at
- created_at
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, Iterable, List

import sys

KB = Path(__file__).resolve().parents[1]
RESOLVER_DIR = KB / "resolver"
GRAPH_DIR = KB / "PropertyGraph"
IMPORT_DIR = GRAPH_DIR / "incoming_bridge_datasets"
IMPORT_DIR.mkdir(parents=True, exist_ok=True)
REPORT_MD = GRAPH_DIR / "EXTERNAL_BRIDGE_IMPORT_REPORT.md"
TEMPLATE_CSV = GRAPH_DIR / "external_bridge_dataset_template.csv"

sys.path.append(str(RESOLVER_DIR))
import bridge_data_layer as bridge_layer  # noqa: E402

sys.path.append(str(GRAPH_DIR))
import build_property_graph as graph_builder  # noqa: E402


def iter_csv(path: Path) -> Iterable[Dict[str, object]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        yield from csv.DictReader(handle)


def iter_json(path: Path) -> Iterable[Dict[str, object]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        for row in data:
            if isinstance(row, dict):
                yield row


def iter_xlsx(path: Path) -> Iterable[Dict[str, object]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            headers = None
            for row in rows:
                vals = [str(v).strip() if v is not None else "" for v in row]
                if not headers:
                    if sum(bool(v) for v in vals) < 2:
                        continue
                    headers = vals
                    continue
                if not any(vals):
                    continue
                yield {headers[i]: vals[i] if i < len(vals) else "" for i in range(len(headers))}
    finally:
        wb.close()


def load_rows(path: Path) -> List[Dict[str, object]]:
    if path.suffix.lower() == ".csv":
        return list(iter_csv(path))
    if path.suffix.lower() == ".json":
        return list(iter_json(path))
    if path.suffix.lower() == ".xlsx":
        return list(iter_xlsx(path))
    raise ValueError(f"Unsupported file type: {path}")


def import_file(path: Path) -> Dict[str, object]:
    normalized_rows = [bridge_layer.normalize_bridge_row(raw) for raw in load_rows(path)]
    import_counts = bridge_layer.import_bridge_rows([path])
    bridge_layer.export_bridge_csv()
    graph_summary = graph_builder.build()
    counts = {"exact_bridge": 0, "partial_bridge": 0, "candidate_bridge": 0, "invalid_bridge": 0}
    for classification, value in import_counts.items():
        if classification in counts:
            counts[classification] = value
    report_lines = [
        "# External Bridge Import Report",
        "",
        f"- input_file: `{path}`",
        f"- imported_rows: `{len(normalized_rows)}`",
        f"- exact_bridge: `{counts.get('exact_bridge', 0)}`",
        f"- partial_bridge: `{counts.get('partial_bridge', 0)}`",
        f"- candidate_bridge: `{counts.get('candidate_bridge', 0)}`",
        f"- invalid_bridge: `{counts.get('invalid_bridge', 0)}`",
        "",
        "## Graph Summary After Import",
        f"- canonical_properties: `{graph_summary['canonical_properties']}`",
        f"- bridge_rows: `{graph_summary['bridge_rows']}`",
        f"- verification_queue: `{graph_summary['verification_queue']}`",
    ]
    REPORT_MD.write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    return {
        "input_file": str(path),
        "imported_rows": len(normalized_rows),
        "classification_counts": counts,
        "graph_summary": graph_summary,
    }


def ensure_template() -> None:
    if TEMPLATE_CSV.exists():
        return
    with TEMPLATE_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "source_platform",
                "listing_url",
                "listing_id",
                "broker_reference",
                "permit_number",
                "property_number",
                "plot_number",
                "building_name",
                "unit_number",
                "canonical_property_id",
                "resolver_record_id",
                "confidence",
                "source_updated_at",
                "created_at",
                "source_file",
                "source_sheet",
                "row_number",
            ]
        )


def main() -> None:
    ensure_template()
    parser = argparse.ArgumentParser()
    parser.add_argument("path", nargs="?", default="")
    args = parser.parse_args()
    if not args.path:
        print(json.dumps({"template": str(TEMPLATE_CSV), "incoming_dir": str(IMPORT_DIR)}, ensure_ascii=False, indent=2))
        return
    result = import_file(Path(args.path))
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
