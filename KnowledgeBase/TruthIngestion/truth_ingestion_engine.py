#!/usr/bin/env python3
"""AIOS Truth Ingestion Engine.

Index verified bridge sources into the existing AIOS bridge/property graph layer.
This script does not change Unit Finder or matching logic. It feeds normalized
truth into the current bridge table and rebuilds downstream graph artifacts.
"""

from __future__ import annotations

import csv
import json
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Dict, Iterable, List

from openpyxl import load_workbook

KB = Path("/Users/hassanka/Downloads/AIOS/KnowledgeBase")
BASE_DIR = KB / "TruthIngestion"
INCOMING_GOOGLE = BASE_DIR / "incoming_sources" / "google_drive"
INCOMING_LOCAL = BASE_DIR / "incoming_sources" / "local"
OUT_DIR = BASE_DIR / "outputs"
OUT_DIR.mkdir(parents=True, exist_ok=True)

NORMALIZED_CSV = OUT_DIR / "truth_ingestion_normalized_bridge_rows.csv"
SOURCE_STATUS_CSV = OUT_DIR / "truth_ingestion_source_status.csv"
REPORT_MD = OUT_DIR / "TRUTH_INGESTION_REPORT.md"
SUMMARY_JSON = OUT_DIR / "truth_ingestion_summary.json"
TEMP_IMPORT_CSV = OUT_DIR / "truth_ingestion_import_payload.csv"

RESOLVER_DIR = KB / "resolver"
PROPERTY_GRAPH_DIR = KB / "PropertyGraph"
BRIDGE_ENGINE_DIR = KB / "BridgeEngine"
INDEX_CSV = RESOLVER_DIR / "unit_resolver_index.csv"
PROPERTY_GRAPH_SUMMARY = PROPERTY_GRAPH_DIR / "property_graph_summary.json"
BRIDGE_ENGINE_MANIFEST = BRIDGE_ENGINE_DIR / "bridge_engine_manifest.json"

import sys

sys.path.append(str(RESOLVER_DIR))
import bridge_data_layer as bridge_layer  # noqa: E402
import listing_similarity_matcher as matcher  # noqa: E402

sys.path.append(str(PROPERTY_GRAPH_DIR))
import build_property_graph as graph_builder  # noqa: E402

URL_RE = re.compile(r"(?:https?://)?(?:www\.)?(?:propertyfinder\.ae|bayut\.com|dubizzle\.com)[^\s\"'<>)\]]+", re.I)
PORTAL_TOKEN_RE = re.compile(r"(property\s*finder|bayut|dubizzle)", re.I)


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def low(value: object) -> str:
    return norm(value).lower()


def digits_only(value: object) -> str:
    text = re.sub(r"\D", "", norm(value))
    return text


def read_json(path: Path) -> Dict[str, object]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def pick_first(*values: object) -> str:
    for value in values:
        text = norm(value)
        if text:
            return text
    return ""


def extract_url(*values: object) -> str:
    for value in values:
        text = norm(value)
        if not text:
            continue
        m = URL_RE.search(text)
        if m:
            return bridge_layer.clean_url(m.group(0))
    return ""


def detect_portal(*values: object) -> str:
    joined = " ".join(norm(v) for v in values if norm(v))
    m = PORTAL_TOKEN_RE.search(joined)
    if not m:
        return ""
    token = low(m.group(1))
    if "property" in token:
        return "propertyfinder.ae"
    if "bayut" in token:
        return "bayut.com"
    if "dubizzle" in token:
        return "dubizzle.com"
    return ""


def price_clue(value: object) -> str:
    text = norm(value).replace(",", "")
    if not text or text == "-":
        return ""
    return text


def size_clue(value: object) -> str:
    text = norm(value).replace(",", "")
    if not text or text == "-":
        return ""
    return text


def split_location_text(value: str) -> Dict[str, str]:
    text = norm(value)
    if not text:
        return {"community": "", "building_name": "", "project_name": ""}
    parts = [part.strip() for part in text.split(",") if norm(part)]
    building_name = parts[0] if parts else ""
    community = parts[-1] if len(parts) > 1 else ""
    project_name = parts[1] if len(parts) > 2 else (parts[0] if len(parts) == 2 else "")
    return {
        "community": community,
        "building_name": building_name,
        "project_name": project_name,
    }


def location_hierarchy(row: Dict[str, object]) -> Dict[str, str]:
    loc1 = norm(row.get("Location 1", ""))
    loc2 = norm(row.get("Location 2", ""))
    loc3 = norm(row.get("Location 3", ""))
    loc4 = norm(row.get("Location 4", ""))
    loc5 = norm(row.get("Location 5", ""))
    loc6 = norm(row.get("Location 6", ""))
    community = pick_first(loc3, loc2, loc1)
    project_name = pick_first(loc5, loc4, loc6)
    building_name = pick_first(loc6, loc5, loc4)
    return {
        "community": community,
        "project_name": project_name,
        "building_name": building_name,
    }


def rows_from_standard_sheet(path: Path, sheet_name: str) -> Iterable[Dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    try:
        rows = ws.iter_rows(values_only=True)
        headers = None
        for raw in rows:
            values = [str(v).strip() if v is not None else "" for v in raw]
            if headers is None:
                if sum(bool(v) for v in values) < 2:
                    continue
                headers = values
                continue
            if headers and headers[0].endswith("Listings") and values and values[0] == "ID":
                headers = values
                continue
            if not any(values):
                continue
            row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
            if low(row.get("ID", "")) == "id":
                continue
            row["_sheet_name"] = sheet_name
            yield row
    finally:
        wb.close()


def rows_from_simple_sheet(path: Path, sheet_name: str) -> Iterable[Dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    try:
        rows = ws.iter_rows(values_only=True)
        headers = None
        for raw in rows:
            values = [str(v).strip() if v is not None else "" for v in raw]
            if headers is None:
                if sum(bool(v) for v in values) < 2:
                    continue
                headers = values
                continue
            if not any(values):
                continue
            row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
            row["_sheet_name"] = sheet_name
            yield row
    finally:
        wb.close()


@dataclass
class SourceSpec:
    name: str
    path: Path
    loader: Callable[[Path], List[Dict[str, str]]]
    priority: int


def load_resolver_records() -> List[Dict[str, str]]:
    with INDEX_CSV.open("r", encoding="utf-8", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


class ResolverMatcher:
    def __init__(self, records: List[Dict[str, str]]) -> None:
        self.records = records
        self.by_property = defaultdict(list)
        self.by_permit = defaultdict(list)
        self.by_plot = defaultdict(list)
        self.by_listing_id = defaultdict(list)
        self.by_building_unit = defaultdict(list)
        for row in records:
            if digits_only(row.get("property_number", "")):
                self.by_property[digits_only(row.get("property_number", ""))].append(row)
            if digits_only(row.get("permit_number", "")):
                self.by_permit[digits_only(row.get("permit_number", ""))].append(row)
            if digits_only(row.get("plot_number", "")):
                self.by_plot[digits_only(row.get("plot_number", ""))].append(row)
            if norm(row.get("listing_id", "")):
                self.by_listing_id[low(row.get("listing_id", ""))].append(row)
            area = matcher.canonical_area(row.get("area", ""))
            building = pick_first(matcher.canonical_building(row.get("building", "")), matcher.canonical_project(row.get("project", "")))
            unit = matcher.normalize_unit(row.get("unit", ""))
            if area and building and unit:
                self.by_building_unit[(area, building, unit)].append(row)
            elif building and unit:
                self.by_building_unit[("", building, unit)].append(row)

    def attach(self, bridge_row: Dict[str, str]) -> Dict[str, str]:
        candidates: List[Dict[str, str]] = []
        property_digits = digits_only(bridge_row.get("property_number", ""))
        permit_digits = digits_only(bridge_row.get("permit_number", ""))
        plot_digits = digits_only(bridge_row.get("plot_number", ""))
        listing_id = low(bridge_row.get("listing_id", ""))
        area = matcher.canonical_area(bridge_row.get("community", ""))
        building = pick_first(
            matcher.canonical_building(bridge_row.get("building_name", "")),
            matcher.canonical_project(bridge_row.get("project_name", "")),
        )
        unit = matcher.normalize_unit(bridge_row.get("unit_number", ""))
        if property_digits:
            candidates = self.by_property.get(property_digits, [])
        if not candidates and permit_digits:
            candidates = self.by_permit.get(permit_digits, [])
        if not candidates and plot_digits:
            candidates = self.by_plot.get(plot_digits, [])
        if not candidates and listing_id:
            candidates = self.by_listing_id.get(listing_id, [])
        if not candidates and building and unit:
            candidates = self.by_building_unit.get((area, building, unit), []) or self.by_building_unit.get(("", building, unit), [])
        if len(candidates) == 1:
            rec = candidates[0]
            bridge_row["resolver_record_id"] = norm(rec.get("resolver_record_id", ""))
            bridge_row["canonical_property_id"] = f"CPID-{bridge_row['resolver_record_id']}" if bridge_row["resolver_record_id"] else bridge_row.get("canonical_property_id", "")
            bridge_row["owner_contact_available"] = "YES" if low(rec.get("owner_contact_available", "")) == "yes" else "NO"
        return bridge_row


def normalize_source_platform(source_platform: str) -> str:
    platform = low(source_platform)
    if "propertyfinder" in platform or "property finder" in platform:
        return "propertyfinder.ae"
    if "bayut" in platform:
        return "bayut.com"
    if "dubizzle" in platform:
        return "dubizzle.com"
    return norm(source_platform)


def build_bridge_row(raw: Dict[str, object], source_file: str, source_sheet: str, row_number: int, base: Dict[str, object]) -> Dict[str, str]:
    row = {
        "source_platform": normalize_source_platform(base.get("source_platform", "")),
        "listing_url": norm(base.get("listing_url", "")),
        "listing_id": norm(base.get("listing_id", "")),
        "broker_reference": norm(base.get("broker_reference", "")),
        "permit_number": norm(base.get("permit_number", "")),
        "property_number": norm(base.get("property_number", "")),
        "plot_number": norm(base.get("plot_number", "")),
        "building_name": norm(base.get("building_name", "")),
        "unit_number": norm(base.get("unit_number", "")),
        "community": norm(base.get("community", "")),
        "project_name": norm(base.get("project_name", "")),
        "resolver_record_id": norm(base.get("resolver_record_id", "")),
        "canonical_property_id": norm(base.get("canonical_property_id", "")),
        "confidence": norm(base.get("confidence", "")),
        "source_updated_at": norm(base.get("source_updated_at", "")),
        "created_at": now_iso(),
        "size_clues": size_clue(base.get("size_clues", "")),
        "price_clues": price_clue(base.get("price_clues", "")),
        "owner_contact_available": "YES" if low(base.get("owner_contact_available", "")) == "yes" else "NO",
        "source_file": source_file,
        "source_sheet": source_sheet,
        "row_number": str(row_number),
        "raw_payload_json": json.dumps(raw, ensure_ascii=False),
    }
    return row


def file_updated_at(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat()


def load_active_listings(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for idx, raw in enumerate(rows_from_simple_sheet(path, "Sale Listings"), start=2):
        loc = location_hierarchy(raw)
        rows.append(
            build_bridge_row(
                raw,
                path.name,
                raw.get("_sheet_name", "Sale Listings"),
                idx,
                {
                    "source_platform": detect_portal(raw.get("Portals", "")),
                    "listing_id": raw.get("Listing ID", ""),
                    "broker_reference": raw.get("Reference No.", ""),
                    "permit_number": raw.get("Permit Number", ""),
                    "building_name": loc["building_name"],
                    "unit_number": raw.get("Unit Number", ""),
                    "community": loc["community"],
                    "project_name": loc["project_name"],
                    "source_updated_at": file_updated_at(path),
                    "size_clues": raw.get("Area (sqft)", ""),
                    "price_clues": raw.get("Price (AED)", ""),
                },
            )
        )
    return rows


def load_listings_workbook(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for idx, raw in enumerate(rows_from_standard_sheet(path, "Sale"), start=3):
        loc = location_hierarchy(raw)
        rows.append(
            build_bridge_row(
                raw,
                path.name,
                raw.get("_sheet_name", "Sale"),
                idx,
                {
                    "source_platform": detect_portal(raw.get("Portals", ""), raw.get("Source", "")),
                    "listing_id": raw.get("ID", ""),
                    "broker_reference": pick_first(raw.get("Reference", ""), raw.get("Transaction Number", "")),
                    "permit_number": raw.get("Permit Number", ""),
                    "building_name": loc["building_name"],
                    "unit_number": raw.get("Unit No", ""),
                    "community": loc["community"],
                    "project_name": loc["project_name"],
                    "source_updated_at": file_updated_at(path),
                    "size_clues": raw.get("Area", ""),
                    "price_clues": raw.get("Original Price", ""),
                },
            )
        )
    return rows


def load_bayut_overall_lead(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for sheet_name in ("PHONE", "WHATS APP ", "EMAIL"):
        for idx, raw in enumerate(rows_from_simple_sheet(path, sheet_name), start=2):
            listing_url = extract_url(raw.get("Property Link in Message", ""), raw.get("Message from Lead", ""), raw.get("Message", ""))
            rows.append(
                build_bridge_row(
                    raw,
                    path.name,
                    raw.get("_sheet_name", sheet_name),
                    idx,
                    {
                        "source_platform": detect_portal(listing_url, raw.get("Enquiry From", ""), "bayut"),
                        "listing_url": listing_url,
                        "listing_id": raw.get("Listing ID", ""),
                        "broker_reference": raw.get("Reference No.", ""),
                        "building_name": raw.get("Sub Location", ""),
                        "community": raw.get("Location", ""),
                        "project_name": raw.get("Sub Location", ""),
                        "source_updated_at": file_updated_at(path),
                        "price_clues": raw.get("Price (AED)", ""),
                    },
                )
            )
    return rows


def load_secondary_links(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for idx, raw in enumerate(rows_from_simple_sheet(path, "Sheet1"), start=2):
        listing_url = extract_url(raw.get("Property Link", ""))
        loc = split_location_text(raw.get("Location", ""))
        rows.append(
            build_bridge_row(
                raw,
                path.name,
                raw.get("_sheet_name", "Sheet1"),
                idx,
                {
                    "source_platform": detect_portal(listing_url),
                    "listing_url": listing_url,
                    "building_name": loc["building_name"],
                    "community": loc["community"],
                    "project_name": loc["project_name"],
                    "source_updated_at": file_updated_at(path),
                    "size_clues": raw.get("Total SQFT Area", ""),
                    "price_clues": raw.get("Selling Price", ""),
                },
            )
        )
    return rows


def load_dubai_brokers(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            raw_rows = ws.iter_rows(values_only=True)
            headers = None
            for idx, raw in enumerate(raw_rows, start=1):
                values = [str(v).strip() if v is not None else "" for v in raw]
                if headers is None:
                    if sum(bool(v) for v in values) < 2:
                        continue
                    headers = values
                    continue
                if not any(values):
                    continue
                row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
                row["_sheet_name"] = ws.title
                rows.append(
                    build_bridge_row(
                        row,
                        path.name,
                        ws.title,
                        idx,
                        {
                            "building_name": pick_first(row.get("Project", ""), row.get("Location/View", "")),
                            "project_name": row.get("Project", ""),
                            "community": "",
                            "unit_number": pick_first(row.get("Unit Number", ""), row.get("Flat No.", "")),
                            "source_updated_at": file_updated_at(path),
                            "size_clues": pick_first(row.get("Saleable Area (Sq.ft)", ""), row.get("Saleable Area Sq.F", "")),
                            "price_clues": pick_first(row.get("Price", ""), row.get("Purchase Price", "")),
                        },
                    )
                )
    finally:
        wb.close()
    return rows


def load_bitrix_messages(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        for idx, raw in enumerate(csv.DictReader(handle), start=2):
            listing_url = extract_url(raw.get("ZTEXT", ""), raw.get("ZTITLE", ""), raw.get("copied_attachment_path", ""))
            if not listing_url:
                continue
            rows.append(
                build_bridge_row(
                    raw,
                    path.name,
                    "messages",
                    idx,
                    {
                        "source_platform": detect_portal(listing_url),
                        "listing_url": listing_url,
                        "source_updated_at": file_updated_at(path),
                    },
                )
            )
    return rows


def load_runtime_memory_manifest(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    text = path.read_text(encoding="utf-8", errors="ignore")
    for idx, match in enumerate(URL_RE.finditer(text), start=1):
        rows.append(
            build_bridge_row(
                {"matched_url": match.group(0)},
                path.name,
                "manifest",
                idx,
                {
                    "source_platform": detect_portal(match.group(0)),
                    "listing_url": bridge_layer.clean_url(match.group(0)),
                    "source_updated_at": file_updated_at(path),
                },
            )
        )
    return rows


def write_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def count_bridge_fields(rows: List[Dict[str, str]]) -> Dict[str, int]:
    counters = Counter()
    for row in rows:
        if norm(row.get("listing_url", "")):
            counters["listing_urls"] += 1
        if norm(row.get("listing_id", "")):
            counters["listing_ids"] += 1
        if norm(row.get("broker_reference", "")):
            counters["broker_references"] += 1
        if norm(row.get("permit_number", "")):
            counters["permit_numbers"] += 1
        if norm(row.get("property_number", "")):
            counters["property_numbers"] += 1
        if norm(row.get("plot_number", "")):
            counters["plot_numbers"] += 1
        if norm(row.get("building_name", "")):
            counters["building_names"] += 1
        if norm(row.get("unit_number", "")):
            counters["unit_numbers"] += 1
        if norm(row.get("resolver_record_id", "")):
            counters["resolver_links"] += 1
    return dict(counters)


def classify_truth_bridge(after_summary: Dict[str, object], after_manifest: Dict[str, object], imported_rows: int) -> str:
    bridge_rows = int(after_summary.get("bridge_rows", 0) or 0)
    public_listing_cpid_count = int(after_summary.get("public_listing_cpid_count", 0) or 0)
    exact_bridge = int(after_summary.get("bridge_status_counts", {}).get("exact_bridge", 0) or 0)
    url_records = int(after_manifest.get("current_data_counts", {}).get("url_records", 0) or 0)
    if imported_rows <= 0:
        return "FAILED"
    if exact_bridge > 50 and public_listing_cpid_count > 10 and url_records > 300:
        return "LIVE"
    return "PARTIAL"


def run_bridge_engine_refresh() -> None:
    subprocess.run(["python3", str(BRIDGE_ENGINE_DIR / "bridge_engine.py")], check=True)


def build_report(source_status_rows: List[Dict[str, str]], before_summary: Dict[str, object], after_summary: Dict[str, object], before_manifest: Dict[str, object], after_manifest: Dict[str, object], imported_rows: int, classification: str) -> None:
    remaining = [row for row in source_status_rows if row["imported_rows"] == "0"]
    lines = [
        "# AIOS Truth Ingestion Report",
        "",
        f"- generated_at: `{now_iso()}`",
        f"- truth_bridge_classification: `{classification}`",
        f"- imported_rows: `{imported_rows}`",
        "",
        "## Coverage Delta",
        "",
        f"- bridge_rows_before: `{before_summary.get('bridge_rows', 0)}`",
        f"- bridge_rows_after: `{after_summary.get('bridge_rows', 0)}`",
        f"- exact_bridge_before: `{before_summary.get('bridge_status_counts', {}).get('exact_bridge', 0)}`",
        f"- exact_bridge_after: `{after_summary.get('bridge_status_counts', {}).get('exact_bridge', 0)}`",
        f"- public_listing_cpid_before: `{before_summary.get('public_listing_cpid_count', 0)}`",
        f"- public_listing_cpid_after: `{after_summary.get('public_listing_cpid_count', 0)}`",
        f"- url_records_before: `{before_manifest.get('current_data_counts', {}).get('url_records', 0)}`",
        f"- url_records_after: `{after_manifest.get('current_data_counts', {}).get('url_records', 0)}`",
        f"- listing_reference_before: `{before_manifest.get('current_data_counts', {}).get('listing_reference_records', 0)}`",
        f"- listing_reference_after: `{after_manifest.get('current_data_counts', {}).get('listing_reference_records', 0)}`",
        f"- permit_records_before: `{before_manifest.get('current_data_counts', {}).get('permit_records', 0)}`",
        f"- permit_records_after: `{after_manifest.get('current_data_counts', {}).get('permit_records', 0)}`",
        f"- property_number_before: `{before_manifest.get('current_data_counts', {}).get('property_number_records', 0)}`",
        f"- property_number_after: `{after_manifest.get('current_data_counts', {}).get('property_number_records', 0)}`",
        "",
        "## Source Status",
        "",
        "| priority | source | available | imported_rows | exact_bridge | partial_bridge | candidate_bridge | invalid_bridge | resolver_links | notes |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in source_status_rows:
        lines.append(
            f"| {row['priority']} | {row['source']} | {row['available']} | {row['imported_rows']} | {row['exact_bridge']} | {row['partial_bridge']} | {row['candidate_bridge']} | {row['invalid_bridge']} | {row['resolver_links']} | {row['notes']} |"
        )
    lines.extend(
        [
            "",
            "## Remaining Unindexed Truth",
            "",
        ]
    )
    if remaining:
        for row in remaining:
            lines.append(f"- `{row['source']}`: {row['notes']}")
    else:
        lines.append("- No zero-row source remained in the declared priority set.")
    lines.extend(
        [
            "",
            "## Evidence",
            "",
            f"- normalized_rows_csv: `{NORMALIZED_CSV}`",
            f"- source_status_csv: `{SOURCE_STATUS_CSV}`",
            f"- summary_json: `{SUMMARY_JSON}`",
            f"- bridge_export_csv: `{RESOLVER_DIR / 'bridge_records_export.csv'}`",
            f"- property_graph_summary: `{PROPERTY_GRAPH_SUMMARY}`",
            f"- bridge_engine_manifest: `{BRIDGE_ENGINE_MANIFEST}`",
        ]
    )
    REPORT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    before_summary = read_json(PROPERTY_GRAPH_SUMMARY)
    before_manifest = read_json(BRIDGE_ENGINE_MANIFEST)
    resolver_matcher = ResolverMatcher(load_resolver_records())

    specs = [
        SourceSpec("Active Listings.xlsx", INCOMING_GOOGLE / "Active Listings.xlsx", load_active_listings, 1),
        SourceSpec("Listings.xlsx", INCOMING_GOOGLE / "Listings.xlsx", load_listings_workbook, 2),
        SourceSpec("Listings (1).xlsx", INCOMING_GOOGLE / "Listings (1).xlsx", load_listings_workbook, 3),
        SourceSpec("BAYUT OVER ALL LEAD.xlsx", INCOMING_GOOGLE / "BAYUT OVER ALL LEAD.xlsx", load_bayut_overall_lead, 4),
        SourceSpec("Secondary Listings LINKS EXCEL.xlsx", INCOMING_GOOGLE / "Secondary Listings LINKS EXCEL.xlsx", load_secondary_links, 5),
        SourceSpec("Dubai_brokers", INCOMING_LOCAL / "Dubai_brokers__Project__5d39d16462.xlsx", load_dubai_brokers, 6),
        SourceSpec("Bitrix24/raw", INCOMING_LOCAL / "messages.csv", load_bitrix_messages, 7),
        SourceSpec("Runtime Memory Ledger", INCOMING_LOCAL / "9645b5021d6c__OPENAI_MEMORY_ARCHIVE_MANIFEST.txt", load_runtime_memory_manifest, 8),
    ]

    all_rows: List[Dict[str, str]] = []
    source_status_rows: List[Dict[str, str]] = []
    for spec in specs:
        raw_rows = spec.loader(spec.path) if spec.path.exists() else []
        attached_rows = [resolver_matcher.attach(bridge_layer.normalize_bridge_row(row)) for row in raw_rows]
        counts = Counter(row.get("bridge_classification", "invalid_bridge") for row in attached_rows)
        field_counts = count_bridge_fields(attached_rows)
        all_rows.extend(attached_rows)
        notes = []
        if not spec.path.exists():
            notes.append("source file not available")
        elif not attached_rows:
            notes.append("available but no bridge-usable rows detected")
        else:
            if field_counts.get("listing_urls", 0):
                notes.append(f"urls={field_counts['listing_urls']}")
            if field_counts.get("listing_ids", 0):
                notes.append(f"listing_ids={field_counts['listing_ids']}")
            if field_counts.get("broker_references", 0):
                notes.append(f"broker_refs={field_counts['broker_references']}")
            if field_counts.get("permit_numbers", 0):
                notes.append(f"permits={field_counts['permit_numbers']}")
            if field_counts.get("unit_numbers", 0):
                notes.append(f"units={field_counts['unit_numbers']}")
            if field_counts.get("resolver_links", 0):
                notes.append(f"resolver_links={field_counts['resolver_links']}")
        source_status_rows.append(
            {
                "priority": str(spec.priority),
                "source": spec.name,
                "available": "YES" if spec.path.exists() else "NO",
                "imported_rows": str(len(attached_rows)),
                "exact_bridge": str(counts.get("exact_bridge", 0)),
                "partial_bridge": str(counts.get("partial_bridge", 0)),
                "candidate_bridge": str(counts.get("candidate_bridge", 0)),
                "invalid_bridge": str(counts.get("invalid_bridge", 0)),
                "resolver_links": str(field_counts.get("resolver_links", 0)),
                "notes": "; ".join(notes),
            }
        )

    deduped_rows: Dict[str, Dict[str, str]] = {}
    for row in all_rows:
        bridge_id = bridge_layer.bridge_id_for_row(row)
        row["bridge_id"] = bridge_id
        deduped_rows[bridge_id] = row
    final_rows = list(deduped_rows.values())

    write_csv(NORMALIZED_CSV, final_rows)
    write_csv(SOURCE_STATUS_CSV, source_status_rows)
    write_csv(TEMP_IMPORT_CSV, final_rows)

    bridge_layer.import_bridge_rows([TEMP_IMPORT_CSV], reset=False)
    bridge_layer.export_bridge_csv()
    after_summary = graph_builder.build()
    run_bridge_engine_refresh()
    after_manifest = read_json(BRIDGE_ENGINE_MANIFEST)

    classification = classify_truth_bridge(after_summary, after_manifest, len(final_rows))
    build_report(source_status_rows, before_summary, after_summary, before_manifest, after_manifest, len(final_rows), classification)

    summary = {
        "generated_at": now_iso(),
        "classification": classification,
        "indexed_sources": [row["source"] for row in source_status_rows if row["available"] == "YES"],
        "rows_imported": len(final_rows),
        "source_status_rows": source_status_rows,
        "before_summary": before_summary,
        "after_summary": after_summary,
        "before_manifest": before_manifest,
        "after_manifest": after_manifest,
        "evidence": {
            "normalized_rows_csv": str(NORMALIZED_CSV),
            "source_status_csv": str(SOURCE_STATUS_CSV),
            "report_md": str(REPORT_MD),
            "bridge_export_csv": str(RESOLVER_DIR / "bridge_records_export.csv"),
            "property_graph_summary": str(PROPERTY_GRAPH_SUMMARY),
            "bridge_engine_manifest": str(BRIDGE_ENGINE_MANIFEST),
        },
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
