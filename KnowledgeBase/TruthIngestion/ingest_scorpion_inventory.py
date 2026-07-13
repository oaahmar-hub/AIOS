#!/usr/bin/env python3
"""Ingest the Scorpion Property shared inventory sheet into the resolver corpus.

Source: "Inventory (Sale/Rent)" Google Sheet (owner scorpionpropertylist@gmail.com),
shared with o.a.ahmar@gmail.com since 2024 — an authorized broker inventory feed.
A dated snapshot (.xlsx) is stored next to this script as provenance.

Rows look like ``EB Seapoint T2 - 1501`` — a building name plus a real unit
number. This importer:

- parses ``building - unit`` from the Property Name column (only when the
  trailing token is a plausible unit: digits or letter+digits; "Full Floor",
  "Land" and blank rows are skipped, never guessed);
- normalizes area/bedrooms/price/size/developer from the sheet columns;
- de-duplicates by (building, unit), preferring the row with a sale price;
- appends the records to ``unit_resolver_index.csv`` **and** the
  ``resolver_records`` table with ``source_file`` provenance so the whole
  batch can be identified or removed later.

Deterministic and idempotent: record ids are content hashes; existing ids are
skipped on re-run. No listing URLs are invented — this is inventory-side
enrichment only.
"""

from __future__ import annotations

import csv
import hashlib
import re
import sqlite3
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

HERE = Path(__file__).resolve().parent
KB = HERE.parent
RESOLVER_DIR = KB / "resolver"
INDEX_CSV = RESOLVER_DIR / "unit_resolver_index.csv"
DB_PATH = RESOLVER_DIR / "unit_resolver_database.resolver"

SNAPSHOT = HERE / "scorpion_inventory_2026-07-07.xlsx"
NORMALIZED_CSV = HERE / "scorpion_inventory_normalized_2026-07-07.csv"

SOURCE_FILE = "Scorpion_Inventory_GDrive_2026-07-07.xlsx"
SOURCE_PATH = "gdrive://1H10g9KNOZbV9hvX21AEMLZ3kKNRrFrUz/Inventory (Sale-Rent) - 07 July 2026"

# Broker shorthand prefixes that mark the master community, not the building.
PREFIXES = {
    "EB": "Beach Front",
    "DE": "Dubai Hills Estate",
    "DC": "Creek Harbour",
    "TV": "Valley",
    "TAG": "Tilal Al Ghaf",
    "MBR": "MBR",
    "TWR": "",
    "DHE": "Dubai Hills",
    "EH": "Emirates Hills",
    "ET": "Emirates Hills",
}

UNIT_RE = re.compile(r"^[a-z]{0,3}[- ]?\d{1,5}[a-z]?$", re.IGNORECASE)
SKIP_UNIT_TOKENS = {"floor", "land", "units", "unit", "plot"}


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip()


def parse_property_name(name: str) -> Optional[Dict[str, str]]:
    """Split ``EB Seapoint T2 - 1501`` into building + unit, or None."""
    text = _norm(name)
    if not text or "full floor" in text.lower() or "half floor" in text.lower():
        return None
    if " - " in text:
        head, _, tail = text.rpartition(" - ")
    else:
        parts = text.rsplit(" ", 1)
        if len(parts) != 2:
            return None
        head, tail = parts
    unit = tail.strip()
    if not unit or not any(ch.isdigit() for ch in unit):
        return None
    if any(tok in unit.lower() for tok in SKIP_UNIT_TOKENS):
        return None
    if not UNIT_RE.fullmatch(unit.replace(" ", "")):
        return None
    building = head.strip(" -")
    tokens = building.split()
    if tokens and tokens[0].upper() in PREFIXES:
        tokens = tokens[1:]
    building = " ".join(tokens).strip()
    if not building or not re.search(r"[a-zA-Z]{3,}", building):
        return None
    return {"building": building, "unit": unit.upper()}


def _clean_number(value: object) -> str:
    text = _norm(value).replace(",", "").replace("-", "")
    try:
        num = float(text)
    except ValueError:
        return ""
    if num <= 0:
        return ""
    return str(int(num)) if num == int(num) else str(num)


def _bedrooms(value: object) -> str:
    text = _norm(value).lower()
    if "studio" in text:
        return "0"
    match = re.search(r"(\d+)", text)
    return match.group(1) if match else ""


def extract_rows() -> List[Dict[str, str]]:
    wb = openpyxl.load_workbook(SNAPSHOT, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows: List[Dict[str, str]] = []
    header: Dict[str, int] = {}
    for raw in sheet.iter_rows(values_only=True):
        cells = [_norm(c) for c in raw]
        if not any(cells):
            continue
        lowered = [c.lower() for c in cells]
        if "property name" in lowered:
            header = {name: idx for idx, name in enumerate(lowered)}
            continue
        if not header:
            continue

        def col(*names: str) -> str:
            for n in names:
                if n in header and header[n] < len(cells):
                    return cells[header[n]]
            return ""

        parsed = parse_property_name(col("property name"))
        if not parsed:
            continue
        rows.append(
            {
                **parsed,
                "area": col("area"),
                "bedrooms": _bedrooms(col("bedroom")),
                "price": _clean_number(col(" price ", "price", " price")),
                "size": _clean_number(col("bua sqft", "plot sqft")),
                "developer": col("developer"),
                "status": col("status"),
            }
        )
    wb.close()

    # De-duplicate by (building, unit); prefer rows carrying a price.
    best: Dict[tuple, Dict[str, str]] = {}
    for row in rows:
        key = (row["building"].lower(), row["unit"].lower())
        current = best.get(key)
        if current is None or (row["price"] and not current["price"]):
            best[key] = row
    return list(best.values())


def record_id(row: Dict[str, str]) -> str:
    basis = f"scorpion2026|{row['building'].lower()}|{row['unit'].lower()}|{row['area'].lower()}"
    return hashlib.sha1(basis.encode("utf-8")).hexdigest()[:16]


def build_record(row: Dict[str, str], fieldnames: List[str]) -> Dict[str, str]:
    record = {name: "" for name in fieldnames}
    record.update(
        {
            "resolver_record_id": record_id(row),
            "area": row["area"],
            "building": row["building"],
            "unit": row["unit"],
            "bedrooms": row["bedrooms"],
            "price": row["price"],
            "size": row["size"],
            "developer": row["developer"],
            "owner_contact_available": "NO",
            "source_file": SOURCE_FILE,
            "source_path": SOURCE_PATH,
            "source_sheet": "Inventory (Sale/Rent)",
            "extracted_from_pdf": "NO",
            "extracted_from_sheet": "YES",
            "extracted_from_text": "NO",
            "extraction_confidence": "85",
            "confidence_score": "85",
            "match_basis": "external_inventory_feed",
        }
    )
    return record


def main() -> None:
    rows = extract_rows()
    with NORMALIZED_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["building", "unit", "area", "bedrooms", "price", "size", "developer", "status"],
        )
        writer.writeheader()
        writer.writerows(rows)

    with INDEX_CSV.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        existing_ids = {r["resolver_record_id"] for r in reader}

    new_records = [
        build_record(row, fieldnames)
        for row in rows
        if record_id(row) not in existing_ids
    ]
    with INDEX_CSV.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writerows(new_records)

    con = sqlite3.connect(DB_PATH)
    try:
        db_cols = [r[1] for r in con.execute("PRAGMA table_info(resolver_records)")]
        db_existing = {r[0] for r in con.execute("select resolver_record_id from resolver_records")}
        inserted = 0
        for row in rows:
            rid = record_id(row)
            if rid in db_existing:
                continue
            record = build_record(row, db_cols)
            cols = ", ".join(record.keys())
            marks = ", ".join("?" for _ in record)
            con.execute(f"insert into resolver_records ({cols}) values ({marks})", list(record.values()))
            inserted += 1
        con.commit()
    finally:
        con.close()

    print(
        {
            "sheet_units_parsed": len(rows),
            "index_rows_appended": len(new_records),
            "db_rows_inserted": inserted,
            "normalized_csv": str(NORMALIZED_CSV.name),
        }
    )


if __name__ == "__main__":
    main()
