#!/usr/bin/env python3
from __future__ import annotations

"""AIOS vector ingestion pipeline.

Adds an unstructured semantic layer beside Property_Master_Database.sqlite.
The structured SQLite ledger remains the source of truth for exact unit,
owner, price, and inventory lookups.
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sqlite3
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook


ROOT = Path("/Users/hassanka/Downloads/AIOS/KnowledgeBase")
RAW_ROOT = ROOT / "Raw"
DEFAULT_SOURCE_ROOT = RAW_ROOT / "CompleteAcquisition_20260621_final"
SQLITE_DB = ROOT / "Property_Master_Database.sqlite"
VECTOR_DIR = ROOT / "VectorDB" / "chroma"
COLLECTION_NAME = "aios_knowledge_chunks"
SUPPORTED_EXTENSIONS = {".xlsx", ".csv", ".pdf", ".txt", ".md"}
MAX_CHARS = 1400
OVERLAP_CHARS = 220
BATCH_SIZE = 250


@dataclass
class ChunkRecord:
    chunk_id: str
    text: str
    metadata: dict


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u200e", "").replace("\u200f", "")).strip()


def canon(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean(value).lower()).strip()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8", errors="ignore")).hexdigest()


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def stable_chunk_id(path: Path, local_id: str, text: str) -> str:
    return sha256_text(f"{path.resolve()}::{local_id}::{sha256_text(text)}")[:48]


def infer_source_group(path: Path, source_root: Path) -> str:
    try:
        rel = path.relative_to(source_root)
        if len(rel.parts) > 1:
            return rel.parts[0]
    except Exception:
        pass
    try:
        rel = path.relative_to(RAW_ROOT)
        return rel.parts[0] if rel.parts else ""
    except Exception:
        return path.parent.name


def header_map(headers: list[str]) -> dict[str, int]:
    aliases = {
        "project": ["project", "project name", "projectname", "projectnameen", "master project", "development"],
        "building": ["buildingnameen", "building name", "building", "tower", "tower name"],
        "unit": ["unitnumber", "unit number", "unit no", "unit", "unit code", "apartment", "plot", "villa"],
        "owner": ["owner", "owner name", "nameen", "buyer name", "seller name", "client name", "customer name"],
    }
    out = {}
    normalized = [canon(h) for h in headers]
    for key, names in aliases.items():
        for wanted in names:
            for i, header in enumerate(normalized):
                if header == wanted:
                    out[key] = i
                    break
            if key in out:
                break
        if key in out:
            continue
        for wanted in names:
            for i, header in enumerate(normalized):
                if wanted not in header:
                    continue
                if key == "owner" and header in {
                    "buildingnameen",
                    "projectnameen",
                    "procedurenameen",
                    "countrynameen",
                    "procedurepartytypenameen",
                }:
                    continue
                if key == "unit" and "plot pre reg" in header:
                    continue
                out[key] = i
                break
            if key in out:
                break
    return out


def enrich_metadata(
    *,
    path: Path,
    source_root: Path,
    sheet: str = "",
    row_number: int | str = "",
    source_group: str = "",
    project: str = "",
    building: str = "",
    unit: str = "",
    owner: str = "",
    page: int | str = "",
) -> dict:
    return {
        "project": clean(project),
        "building": clean(building or project),
        "unit": clean(unit),
        "owner": clean(owner),
        "source_file": str(path.resolve()),
        "source_file_name": path.name,
        "sheet": clean(sheet),
        "row_number": str(row_number or ""),
        "source_group": clean(source_group or infer_source_group(path, source_root)),
        "page": str(page or ""),
        "file_type": path.suffix.lower().lstrip("."),
    }


def semantic_chunks(text: str, max_chars: int = MAX_CHARS, overlap_chars: int = OVERLAP_CHARS) -> Iterator[str]:
    text = re.sub(r"\r\n?", "\n", text or "")
    paragraphs = [clean(p) for p in re.split(r"\n\s*\n|\n(?=[A-Z0-9][^\\n]{0,80}:)", text) if clean(p)]
    if not paragraphs:
        paragraphs = [clean(text)] if clean(text) else []
    buf = ""
    for para in paragraphs:
        if not buf:
            buf = para
        elif len(buf) + len(para) + 2 <= max_chars:
            buf += "\n" + para
        else:
            if buf:
                yield buf
            overlap = buf[-overlap_chars:] if overlap_chars and len(buf) > overlap_chars else ""
            buf = clean((overlap + "\n" + para) if overlap else para)
            while len(buf) > max_chars:
                yield buf[:max_chars]
                buf = clean(buf[max_chars - overlap_chars :])
    if buf:
        yield buf


def iter_xlsx(path: Path, source_root: Path) -> Iterator[ChunkRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    source_group = infer_source_group(path, source_root)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        headers = [clean(v) for v in next(rows, [])]
        hmap = header_map(headers)
        for row_number, row in enumerate(rows, start=2):
            values = [clean(v) for v in row]
            if not any(values):
                continue
            row_pairs = []
            for idx, value in enumerate(values):
                if value:
                    header = headers[idx] if idx < len(headers) and headers[idx] else f"Column {idx + 1}"
                    row_pairs.append(f"{header}: {value}")
            text = f"Source file: {path.name}\nSheet: {ws.title}\nRow: {row_number}\n" + "\n".join(row_pairs)
            project = values[hmap["project"]] if "project" in hmap and hmap["project"] < len(values) else ""
            building = values[hmap["building"]] if "building" in hmap and hmap["building"] < len(values) else project
            unit = values[hmap["unit"]] if "unit" in hmap and hmap["unit"] < len(values) else ""
            owner = values[hmap["owner"]] if "owner" in hmap and hmap["owner"] < len(values) else ""
            metadata = enrich_metadata(
                path=path,
                source_root=source_root,
                sheet=ws.title,
                row_number=row_number,
                source_group=source_group,
                project=project,
                building=building,
                unit=unit,
                owner=owner,
            )
            yield ChunkRecord(stable_chunk_id(path, f"{ws.title}:{row_number}", text), text, metadata)


def iter_csv(path: Path, source_root: Path) -> Iterator[ChunkRecord]:
    source_group = infer_source_group(path, source_root)
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        headers = reader.fieldnames or []
        for row_number, row in enumerate(reader, start=2):
            if not any(clean(v) for v in row.values()):
                continue
            row_text = "\n".join(f"{clean(k)}: {clean(v)}" for k, v in row.items() if clean(v))
            text = f"Source file: {path.name}\nRow: {row_number}\n{row_text}"
            lower = {canon(k): clean(v) for k, v in row.items()}
            metadata = enrich_metadata(
                path=path,
                source_root=source_root,
                row_number=row_number,
                source_group=source_group,
                project=lower.get("project") or lower.get("project name") or lower.get("building", ""),
                building=lower.get("building") or lower.get("tower") or lower.get("project", ""),
                unit=lower.get("unit") or lower.get("unit no") or lower.get("unit number", ""),
                owner=lower.get("owner") or lower.get("owner name") or lower.get("seller", ""),
            )
            yield ChunkRecord(stable_chunk_id(path, str(row_number), text), text, metadata)


def iter_pdf(path: Path, source_root: Path) -> Iterator[ChunkRecord]:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise RuntimeError("pypdf is required for PDF ingestion") from exc
    source_group = infer_source_group(path, source_root)
    reader = PdfReader(str(path))
    for page_num, page in enumerate(reader.pages, start=1):
        text = clean(page.extract_text() or "")
        if not text:
            continue
        for idx, chunk in enumerate(semantic_chunks(text), start=1):
            metadata = enrich_metadata(path=path, source_root=source_root, source_group=source_group, page=page_num)
            yield ChunkRecord(stable_chunk_id(path, f"page-{page_num}:{idx}", chunk), chunk, metadata)


def iter_text(path: Path, source_root: Path) -> Iterator[ChunkRecord]:
    source_group = infer_source_group(path, source_root)
    text = path.read_text(encoding="utf-8", errors="ignore")
    for idx, chunk in enumerate(semantic_chunks(text), start=1):
        metadata = enrich_metadata(path=path, source_root=source_root, source_group=source_group, row_number=idx)
        yield ChunkRecord(stable_chunk_id(path, f"text:{idx}", chunk), chunk, metadata)


def iter_file_chunks(path: Path, source_root: Path) -> Iterator[ChunkRecord]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        yield from iter_xlsx(path, source_root)
    elif ext == ".csv":
        yield from iter_csv(path, source_root)
    elif ext == ".pdf":
        yield from iter_pdf(path, source_root)
    elif ext in {".txt", ".md"}:
        yield from iter_text(path, source_root)


def get_chroma_collection():
    try:
        import chromadb
    except Exception as exc:
        raise RuntimeError("chromadb is required. Install with: python3 -m pip install --user chromadb") from exc
    VECTOR_DIR.mkdir(parents=True, exist_ok=True)
    client = chromadb.PersistentClient(path=str(VECTOR_DIR))
    return client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine", "description": "AIOS semantic knowledge chunks"},
    )


def iter_sqlite_inventory_chunks(limit_rows: int | None = None, offset: int = 0) -> Iterator[ChunkRecord]:
    sql = f"""
        SELECT
            ir.inventory_row_id,
            ir.raw_json,
            ir.unit_ref,
            ir.bedrooms,
            ir.price,
            ir.status,
            ir.inventory_type,
            IFNULL(p.name, '') AS project,
            IFNULL(a.name, '') AS area,
            IFNULL(d.name, '') AS developer,
            IFNULL(pt.name, '') AS property_type,
            IFNULL(f.file_name, '') AS source_file_name,
            IFNULL(f.source_path, '') AS source_file,
            IFNULL(f.source_group, '') AS source_group,
            ir.sheet_name AS sheet,
            ir.row_number
        FROM inventory_rows ir
        LEFT JOIN projects p ON ir.project_id = p.project_id
        LEFT JOIN areas a ON ir.area_id = a.area_id
        LEFT JOIN developers d ON ir.developer_id = d.developer_id
        LEFT JOIN property_types pt ON ir.property_type_id = pt.property_type_id
        LEFT JOIN inventory_files f ON ir.source_id = f.source_id
        ORDER BY ir.inventory_row_id
        LIMIT {int(limit_rows) if limit_rows else -1}
        OFFSET {int(offset)}
    """
    con = sqlite3.connect(f"file:{SQLITE_DB}?mode=ro", uri=True, timeout=10)
    con.row_factory = sqlite3.Row
    try:
        for row in con.execute(sql):
            item = dict(row)
            raw = {}
            try:
                raw = json.loads(item.get("raw_json") or "{}")
            except Exception:
                raw = {}
            source_path = Path(item.get("source_file") or item.get("source_file_name") or "Property_Master_Database.sqlite")
            project = clean(item.get("project") or raw.get("project"))
            building = clean(raw.get("building") or raw.get("BuildingNameEn") or raw.get("building_name") or project)
            unit = clean(item.get("unit_ref") or raw.get("unit_ref") or raw.get("unit") or raw.get("UnitNumber"))
            owner = clean(raw.get("owner") or raw.get("NameEn") or raw.get("name") or "")
            text = "\n".join(
                [
                    "Source: Property_Master_Database.sqlite",
                    f"Inventory Row ID: {item.get('inventory_row_id')}",
                    f"Project: {project}",
                    f"Area: {clean(item.get('area') or raw.get('area'))}",
                    f"Building: {building}",
                    f"Unit: {unit}",
                    f"Owner: {owner}",
                    f"Developer: {clean(item.get('developer') or raw.get('developer'))}",
                    f"Property Type: {clean(item.get('property_type') or raw.get('property_type'))}",
                    f"Bedrooms: {clean(item.get('bedrooms') or raw.get('bedrooms'))}",
                    f"Price: {clean(item.get('price') or raw.get('price'))}",
                    f"Status: {clean(item.get('status') or raw.get('status'))}",
                    f"Inventory Type: {clean(item.get('inventory_type') or raw.get('inventory_type'))}",
                    f"Source File: {clean(item.get('source_file_name'))}",
                    f"Sheet: {clean(item.get('sheet'))}",
                    f"Row Number: {clean(item.get('row_number'))}",
                    "Raw: " + clean(json.dumps(raw, ensure_ascii=False)),
                ]
            )
            metadata = {
                "project": project,
                "building": building,
                "unit": unit,
                "owner": owner,
                "source_file": clean(item.get("source_file") or item.get("source_file_name")),
                "source_file_name": clean(item.get("source_file_name")),
                "sheet": clean(item.get("sheet")),
                "row_number": str(item.get("row_number") or ""),
                "source_group": clean(item.get("source_group")),
                "page": "",
                "file_type": "sqlite_inventory",
                "inventory_row_id": str(item.get("inventory_row_id")),
            }
            yield ChunkRecord(f"sqlite_inventory_{item.get('inventory_row_id')}", text, metadata)
    finally:
        con.close()


def ingest_sqlite_ledger(limit_rows: int | None = None, offset: int = 0) -> dict:
    init_audit_tables()
    collection = get_chroma_collection()
    stats = {"rows_seen": 0, "chunks_upserted": 0}
    ids: list[str] = []
    docs: list[str] = []
    metas: list[dict] = []
    for record in iter_sqlite_inventory_chunks(limit_rows=limit_rows, offset=offset):
        stats["rows_seen"] += 1
        ids.append(record.chunk_id)
        docs.append(record.text)
        metas.append(record.metadata)
        if len(ids) >= BATCH_SIZE:
            collection.upsert(ids=ids, documents=docs, metadatas=metas)
            stats["chunks_upserted"] += len(ids)
            ids, docs, metas = [], [], []
    if ids:
        collection.upsert(ids=ids, documents=docs, metadatas=metas)
        stats["chunks_upserted"] += len(ids)
    con = sqlite3.connect(SQLITE_DB)
    con.execute(
        """
        INSERT INTO vector_ingestion_audit(source_file, file_sha256, chunk_count, last_ingested_ts, status, error)
        VALUES (?, ?, ?, ?, 'ok', '')
        ON CONFLICT(source_file) DO UPDATE SET
            file_sha256=excluded.file_sha256,
            chunk_count=excluded.chunk_count,
            last_ingested_ts=excluded.last_ingested_ts,
            status=excluded.status,
            error=''
        """,
        (
            str(SQLITE_DB.resolve()) + f"#inventory_rows:{offset}:{limit_rows or 'all'}",
            sha256_text(f"{SQLITE_DB.resolve()}:{offset}:{limit_rows or 'all'}"),
            stats["chunks_upserted"],
            int(time.time()),
        ),
    )
    con.commit()
    con.close()
    return stats


def init_audit_tables():
    con = sqlite3.connect(SQLITE_DB)
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS vector_ingestion_audit (
            source_file TEXT PRIMARY KEY,
            file_sha256 TEXT NOT NULL,
            chunk_count INTEGER NOT NULL,
            last_ingested_ts INTEGER NOT NULL,
            status TEXT NOT NULL,
            error TEXT DEFAULT ''
        )
        """
    )
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS retrieval_audit_log (
            audit_id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_ts INTEGER NOT NULL,
            user_query TEXT NOT NULL,
            executed_sql_query TEXT DEFAULT '',
            vector_chunks_retrieved TEXT DEFAULT '[]',
            merged_confidence_level REAL DEFAULT 0,
            sql_result_count INTEGER DEFAULT 0,
            vector_result_count INTEGER DEFAULT 0
        )
        """
    )
    con.commit()
    con.close()


def already_ingested(path: Path, digest: str) -> bool:
    con = sqlite3.connect(SQLITE_DB)
    row = con.execute(
        "SELECT file_sha256, status FROM vector_ingestion_audit WHERE source_file = ?",
        (str(path.resolve()),),
    ).fetchone()
    con.close()
    return bool(row and row[0] == digest and row[1] == "ok")


def update_ingestion_audit(path: Path, digest: str, chunk_count: int, status: str, error: str = ""):
    con = sqlite3.connect(SQLITE_DB)
    con.execute(
        """
        INSERT INTO vector_ingestion_audit(source_file, file_sha256, chunk_count, last_ingested_ts, status, error)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(source_file) DO UPDATE SET
            file_sha256=excluded.file_sha256,
            chunk_count=excluded.chunk_count,
            last_ingested_ts=excluded.last_ingested_ts,
            status=excluded.status,
            error=excluded.error
        """,
        (str(path.resolve()), digest, chunk_count, int(time.time()), status, error[:1000]),
    )
    con.commit()
    con.close()


def discover_files(source_root: Path, limit: int | None = None) -> list[Path]:
    files = [p for p in source_root.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS]
    files = sorted(files)
    return files[:limit] if limit else files


def ingest(source_root: Path = DEFAULT_SOURCE_ROOT, limit_files: int | None = None, force: bool = False) -> dict:
    init_audit_tables()
    collection = get_chroma_collection()
    files = discover_files(source_root, limit_files)
    stats = {"files_seen": len(files), "files_ingested": 0, "files_skipped": 0, "files_failed": 0, "chunks_upserted": 0}
    for path in files:
        digest = file_sha256(path)
        if not force and already_ingested(path, digest):
            stats["files_skipped"] += 1
            continue
        ids: list[str] = []
        docs: list[str] = []
        metas: list[dict] = []
        file_chunk_count = 0
        try:
            for record in iter_file_chunks(path, source_root):
                ids.append(record.chunk_id)
                docs.append(record.text)
                metas.append(record.metadata)
                if len(ids) >= BATCH_SIZE:
                    collection.upsert(ids=ids, documents=docs, metadatas=metas)
                    stats["chunks_upserted"] += len(ids)
                    file_chunk_count += len(ids)
                    ids, docs, metas = [], [], []
            if ids:
                collection.upsert(ids=ids, documents=docs, metadatas=metas)
                stats["chunks_upserted"] += len(ids)
                file_chunk_count += len(ids)
            update_ingestion_audit(path, digest, file_chunk_count, "ok")
            stats["files_ingested"] += 1
        except Exception as exc:
            update_ingestion_audit(path, digest, 0, "failed", str(exc))
            stats["files_failed"] += 1
            print(json.dumps({"file": str(path), "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
    return stats


def main():
    parser = argparse.ArgumentParser(description="Ingest AIOS files into Chroma vector DB")
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT))
    parser.add_argument("--limit-files", type=int, default=None)
    parser.add_argument("--skip-files", action="store_true", help="Do not ingest source files in this run.")
    parser.add_argument("--sqlite-only", action="store_true", help="Only ingest SQLite inventory rows.")
    parser.add_argument("--include-sqlite-ledger", action="store_true")
    parser.add_argument("--limit-sqlite-rows", type=int, default=None)
    parser.add_argument("--sqlite-offset", type=int, default=0)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    result = {"files_seen": 0, "files_ingested": 0, "files_skipped": 0, "files_failed": 0, "chunks_upserted": 0}
    if not args.skip_files and not args.sqlite_only:
        result = ingest(Path(args.source_root), args.limit_files, args.force)

    if args.include_sqlite_ledger or args.sqlite_only:
        result["sqlite_ledger"] = ingest_sqlite_ledger(args.limit_sqlite_rows, args.sqlite_offset)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
